"""
server.py – Backend Flask ERP
Soporta: VIFER (CFDI) y ALESSIA (Pedido SAP)
pip install flask flask-cors pdfplumber openpyxl
"""
import io, re, os
from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

app = Flask(__name__)
CORS(app)

DARK_BLUE="1F3864"; MID_BLUE="2E75B6"; ORANGE="ED7D31"; WHITE="FFFFFF"
LIGHT_GRAY="F2F2F2"; YELLOW="FFFF99"; GREEN="70AD47"; DARK_GRAY="595959"; GOLD="FFC000"

def _thin():
    s=Side(style="thin",color="BFBFBF"); return Border(left=s,right=s,top=s,bottom=s)

def hdr(cell,bg=DARK_BLUE,fg=WHITE,size=10,center=True,bold=True):
    cell.font=Font(name="Arial",bold=bold,color=fg,size=size)
    cell.fill=PatternFill("solid",start_color=bg)
    cell.alignment=Alignment(horizontal="center" if center else "left",vertical="center",wrap_text=True)
    cell.border=_thin()

def dc(cell,bg=None,bold=False,center=False,fmt=None):
    cell.font=Font(name="Arial",size=9,bold=bold)
    cell.alignment=Alignment(horizontal="center" if center else "left",vertical="center")
    cell.border=_thin()
    if bg: cell.fill=PatternFill("solid",start_color=bg)
    if fmt: cell.number_format=fmt

def editable(cell,bg=YELLOW,fmt=None):
    cell.fill=PatternFill("solid",start_color=bg); cell.font=Font(name="Arial",size=9)
    cell.alignment=Alignment(horizontal="right",vertical="center"); cell.border=_thin()
    if fmt: cell.number_format=fmt

# ── Parser VIFER CFDI ─────────────────────────────────────────────────────────
def parse_cfdi(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        lines=[]; ft=""
        for page in pdf.pages:
            t=page.extract_text() or ""; ft+=t+"\n"; lines.extend(t.split("\n"))

    def find(pat,default=""):
        m=re.search(pat,ft,re.MULTILINE)
        if not m: return default
        try: return m.group(1).strip()
        except IndexError: return m.group(0).strip()

    nombre=""; no_pedido=find(r"No\.\s*Pedido\s+(\d+)")
    for i,line in enumerate(lines):
        if re.match(r"Cliente:\s*\d+",line) and i+1<len(lines):
            raw=lines[i+1].strip()
            nombre=re.split(r"\s+(?:Fecha|RFC|R[eé]gimen|Domicilio|CALLE)",raw)[0].strip()
            break
    if not no_pedido:
        for i,line in enumerate(lines):
            if re.match(r"No\.\s*Pedido",line):
                m=re.search(r"(\d{4,6})",line)
                if m: no_pedido=m.group(1)
                elif i>0:
                    m=re.search(r"(\d{4,6})",lines[i-1])
                    if m: no_pedido=m.group(1)
                break

    razon="VILCHES FERRETEROS SA DE CV"
    for line in lines:
        if re.search(r"\b(SA DE CV|S\.A\.|SRL|SC)\b",line,re.IGNORECASE):
            razon=line.strip(); break

    header={"empresa":"vifer","razon_social":razon,
        "rfc_emisor":find(r"RFC\s+(VFE\w+)"),
        "no_factura":find(r"(?:FACTURA|Factura)\s+(A\s*\d+)") or find(r"A\s+(\d{5})"),
        "folio_fiscal":find(r"Folio Fiscal:\s*([A-F0-9\-]+)"),
        "fecha_expedicion":find(r"Expedici[oó]n:\s*(\S+)"),
        "metodo_pago":find(r"M[eé]todo de Pago:\s*(.+)"),
        "forma_pago":find(r"Forma de Pago:\s*(.+)"),
        "uso_cfdi":find(r"Uso CFDI:\s*(.+)"),
        "condiciones_pago":find(r"Condiciones de Pago:\s*(.+)"),
        "no_pedido":no_pedido or "",
        "fecha_vencimiento":find(r"vencimiento\s+(\S+)"),
        "vendedor":find(r"Vendedor:\s*(.+)"),
        "cliente_no":find(r"Cliente:\s*(\d+)"),
        "cliente_nombre":nombre,
        "cliente_rfc":find(r"RFC:\s*([A-Z]{4}\d{6}\w{3})"),
        "cliente_regimen":find(r"R[eé]gimen Fiscal:\s*(.+)"),
        "cliente_domicilio":find(r"(CALLE[^\n]+)")}

    item_re=re.compile(r"^(\d+\.\d{2})\s+(H87|PR)\s+(\d{5,})\s+(.+?)\s+([\d,]+\.\d{2})\s+MXN\s+([\d,]+\.\d{2})\s+MXN\s+([\d,]+\.\d{2})\s+MXN")
    um_id_re=re.compile(r"^(PZA|PR)\s+(\d{4})")
    items=[]; pu=pi=""
    for line in lines:
        m=um_id_re.match(line.strip())
        if m: pu=m.group(1); pi=m.group(2); continue
        m=item_re.match(line.strip())
        if m:
            imp=float(m.group(6).replace(",",""))
            items.append({"no_identif":pi,"descripcion":m.group(4).strip(),
                "cantidad":float(m.group(1)),"um":pu or m.group(2),"clave_sat":m.group(3),
                "precio_unit":float(m.group(5).replace(",","")),"importe":imp,
                "descuento":float(m.group(7).replace(",","")),"iva":round(imp*0.16,2)})
            pu=pi=""
    sub=float((find(r"Subtotal\s+([\d,]+\.\d{2})\s+MXN") or "0").replace(",",""))
    tot=float((find(r"Total\s+([\d,]+\.\d{2})\s+MXN") or "0").replace(",",""))
    letra=find(r"\(([A-Z\u00C1\u00C9\u00CD\u00D3\u00DA\u00D1\s]+\d+/100\s+MN)\)")
    sub=sub or sum(i["importe"] for i in items)
    return {"empresa":"vifer","header":header,"items":items,
            "totals":{"subtotal":sub,"iva":round(sub*0.16,2),"total":tot or round(sub*1.16,2),"letra":letra}}

# ── Parser ALESSIA SAP ────────────────────────────────────────────────────────
def parse_alessia(pdf_bytes):
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        lines=[]; ft=""
        for page in pdf.pages:
            t=page.extract_text() or ""; ft+=t+"\n"; lines.extend(t.split("\n"))

    def find(pat,default=""):
        m=re.search(pat,ft,re.MULTILINE)
        if not m: return default
        try: return m.group(1).strip()
        except IndexError: return m.group(0).strip()

    header={"empresa":"alessia","razon_social":"ALESSIA MOTOPARTES",
        "no_pedido_sap":find(r"Pedido SAP:\s*(\d+)"),
        "no_ecommerce":find(r"E-Commerce:\s*(\d+)"),
        "fecha":find(r"(\d+/\d+/\d{4}\s+\d+:\d+:\d+\w+)"),
        "cliente_codigo":"","cliente_nombre":"","cliente_direccion":"",
        "cliente_ciudad":"","cliente_cp":"",
        "condiciones_pago":find(r"Condiciones de Pago:\s*(.+)"),
        "vendedor":find(r"Empleado Departamento:\s*(.+)"),
        "referencia":find(r"Referencia:\s*\n(\d+)") or find(r"^(\d{5})\s*$",""),
        "bodega":find(r"(BODEGA\s+\d+[^\n]*)")}

    for i,line in enumerate(lines[:15]):
        s=line.strip()
        if re.match(r"CL\d+",s): header["cliente_codigo"]=s.split()[0]
        elif re.match(r"[A-Z\u00C1\u00C9\u00CD\u00D3\u00DA\u00D1]{3,}\s+[A-Z\u00C1\u00C9\u00CD\u00D3\u00DA\u00D1]{3,}",s) and "BODEGA" not in s and "ENVIO" not in s and "ALESSIA" not in s and "Direccion" not in s:
            if not header["cliente_nombre"]: header["cliente_nombre"]=s
        elif re.match(r"CALLE",s): header["cliente_direccion"]=s
        elif re.match(r"\d{5}$",s): header["cliente_cp"]=s
        elif re.match(r"[A-Z\u00C1\u00C9\u00CD\u00D3\u00DA\u00D1\s]{4,}$",s) and "BODEGA" not in s and "ENVIO" not in s and "CENTRO" not in s and "Direccion" not in s and not header["cliente_ciudad"]:
            header["cliente_ciudad"]=s

    item_re=re.compile(r"^(\d+)\s+([A-Za-z0-9][\w\-]*)\s+(.+?)\s+([\d]+\.?\d*)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})$")
    items=[]
    for line in lines:
        m=item_re.match(line.strip())
        if m:
            precio=float(m.group(5).replace(",","")); cant=float(m.group(4)); total=float(m.group(7).replace(",",""))
            items.append({"no_identif":m.group(2),"descripcion":m.group(3).strip(),
                "cantidad":cant,"um":"PZA","clave_sat":"","precio_unit":precio,
                "importe":total,"descuento":float(m.group(6).replace(",","")),"iva":round(total*0.16,2)})

    sub_raw=find(r"Sub Total\s+\$([\d,]+\.\d{2})")
    tot_raw=find(r"^Total\s+\$([\d,]+\.\d{2})") or find(r"Total\s+\$([\d,]+\.\d{2})")
    iva_raw=find(r"Impuesto\s+\$([\d,]+\.\d{2})")
    sub=float(sub_raw.replace(",","")) if sub_raw else sum(i["importe"] for i in items)
    tot=float(tot_raw.replace(",","")) if tot_raw else sub*1.16
    iva=float(iva_raw.replace(",","")) if iva_raw else sub*0.16
    return {"empresa":"alessia","header":header,"items":items,
            "totals":{"subtotal":sub,"iva":iva,"total":tot,"letra":""}}

# ── Generar Excel ─────────────────────────────────────────────────────────────
def build_excel(data,mult_pub,mult_may):
    empresa=data["empresa"]; header=data["header"]; items=data["items"]; totals=data["totals"]
    ACC="C0392B" if empresa=="alessia" else DARK_BLUE
    ACC2="E74C3C" if empresa=="alessia" else MID_BLUE
    seen={}
    for it in items:
        k=it["no_identif"] or it["descripcion"][:20]
        if k not in seen: seen[k]=it
    unique=list(seen.values())
    wb=Workbook()

    # Hoja 1: Catálogo
    ws=wb.active; ws.title="Catalogo Productos"; ws.sheet_properties.tabColor=GREEN
    ws.merge_cells("A1:M1")
    ws["A1"]=f"CATALOGO  -  {header.get('razon_social','').upper()}"
    ws["A1"].font=Font(name="Arial",bold=True,size=13,color=WHITE)
    ws["A1"].fill=PatternFill("solid",start_color=ACC)
    ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=32

    ws.merge_cells("A2:M2")
    ws["A2"]=f"Precio Publico = Costo x {mult_pub}   |   Precio Mayoreo = Costo x {mult_may}   |   Celdas AMARILLAS editables"
    ws["A2"].font=Font(name="Arial",italic=True,size=9,color=DARK_GRAY)
    ws["A2"].fill=PatternFill("solid",start_color="FFF2CC")
    ws["A2"].alignment=Alignment(horizontal="left",vertical="center")
    ws.row_dimensions[2].height=16

    cols=[("A","No.\nArticulo",10,ACC),("B","Descripcion",52,ACC),("C","U.M.",6,ACC),
          ("D","Clave SAT",12,ACC),("E",f"Precio\nCosto",16,ORANGE),
          ("F",f"Precio\nPublico x{mult_pub}",17,"1E7A4A"),
          ("G",f"Precio\nMayoreo x{mult_may}",17,"1F4E8A"),
          ("H","Stock\nActual",12,ORANGE),("I","Stock\nMinimo",12,ORANGE),
          ("J","Ubicacion\nAlmacen",18,ORANGE),("K","Activo\n(Si/No)",10,ORANGE),
          ("L","Ultima\nActual.",16,ORANGE),("M","Notas",28,ORANGE)]
    for cl,title,width,bg in cols:
        c=ws[f"{cl}3"]; c.value=title; hdr(c,bg=bg,size=9)
        ws.column_dimensions[cl].width=width
    ws.row_dimensions[3].height=38
    dv=DataValidation(type="list",formula1='"Si,No"',allow_blank=True); dv.sqref="K4:K2000"
    ws.add_data_validation(dv)

    for idx,prod in enumerate(unique,1):
        r=3+idx; alt=LIGHT_GRAY if idx%2==0 else None
        row=[prod["no_identif"],prod["descripcion"],prod["um"],prod.get("clave_sat",""),
             prod["precio_unit"],None,None,None,None,None,None,None,None]
        for ci,val in enumerate(row,1):
            c=ws.cell(row=r,column=ci)
            if ci==5:
                c.value=val; editable(c,fmt='"$"#,##0.00')
            elif ci==6:
                c.value=f"=E{r}*{mult_pub}"; c.font=Font(name="Arial",size=9,bold=True,color=WHITE)
                c.fill=PatternFill("solid",start_color="1E7A4A")
                c.alignment=Alignment(horizontal="right",vertical="center"); c.border=_thin(); c.number_format='"$"#,##0.00'
            elif ci==7:
                c.value=f"=E{r}*{mult_may}"; c.font=Font(name="Arial",size=9,bold=True,color=WHITE)
                c.fill=PatternFill("solid",start_color="1F4E8A")
                c.alignment=Alignment(horizontal="right",vertical="center"); c.border=_thin(); c.number_format='"$"#,##0.00'
            elif ci in (8,9,10,11,12,13):
                editable(c); c.number_format="#,##0" if ci in (8,9) else "General"
            else:
                c.value=val; dc(c,bg=alt,center=(ci!=2))
    ws.freeze_panes="A4"
    last=3+len(unique)+2; ws.merge_cells(f"A{last}:M{last}")
    ws[f"A{last}"]=f"Publico x{mult_pub}  |  Mayoreo x{mult_may}  |  {header.get('razon_social','')}  |  Doc: {header.get('no_factura',header.get('no_pedido_sap',''))}"
    ws[f"A{last}"].font=Font(name="Arial",italic=True,size=8,color=DARK_GRAY)
    ws[f"A{last}"].fill=PatternFill("solid",start_color="FFF2CC")
    ws[f"A{last}"].alignment=Alignment(horizontal="left",vertical="center")

    # Hoja 2: Encabezado
    ws2=wb.create_sheet("Encabezado"); ws2.sheet_properties.tabColor=ACC2
    ws2.merge_cells("A1:G1"); ws2["A1"]=f"{header.get('razon_social','').upper()}  -  DOCUMENTO"
    hdr(ws2["A1"],bg=ACC,size=13); ws2.row_dimensions[1].height=30
    pairs=([ ("Empresa","ALESSIA MOTOPARTES"),("Pedido SAP",header.get("no_pedido_sap","")),
             ("E-Commerce",header.get("no_ecommerce","")),("Fecha",header.get("fecha","")),
             ("Bodega",header.get("bodega","")),("Referencia",header.get("referencia","")),
             ("Cod. Cliente",header.get("cliente_codigo","")),("Nombre",header.get("cliente_nombre","")),
             ("Direccion",header.get("cliente_direccion","")),("Ciudad",header.get("cliente_ciudad","")),
             ("CP",header.get("cliente_cp","")),("Cond. Pago",header.get("condiciones_pago","")),
             ("Vendedor",header.get("vendedor",""))]
           if empresa=="alessia" else
           [("Empresa",header.get("razon_social","")),("RFC Emisor",header.get("rfc_emisor","")),
            ("No. Factura",header.get("no_factura","")),("Folio Fiscal",header.get("folio_fiscal","")),
            ("Fecha Exp.",header.get("fecha_expedicion","")),("No. Pedido",header.get("no_pedido","")),
            ("Metodo Pago",header.get("metodo_pago","")),("Forma Pago",header.get("forma_pago","")),
            ("Uso CFDI",header.get("uso_cfdi","")),("Cond. Pago",header.get("condiciones_pago","")),
            ("Vendedor",header.get("vendedor","")),("Cliente No.",header.get("cliente_no","")),
            ("Cliente",header.get("cliente_nombre","")),("RFC Cliente",header.get("cliente_rfc","")),
            ("Domicilio",header.get("cliente_domicilio",""))])
    for i,(label,val) in enumerate(pairs):
        r=3+i
        ws2[f"A{r}"]=label; ws2[f"A{r}"].font=Font(name="Arial",bold=True,size=9,color=ACC); ws2[f"A{r}"].border=_thin(); ws2[f"A{r}"].alignment=Alignment(horizontal="left",vertical="center")
        ws2.merge_cells(f"C{r}:G{r}"); ws2[f"C{r}"]=val; ws2[f"C{r}"].font=Font(name="Arial",size=9); ws2[f"C{r}"].border=_thin(); ws2[f"C{r}"].alignment=Alignment(horizontal="left",vertical="center")
        if i%2==0:
            for col in range(1,8): ws2.cell(r,col).fill=PatternFill("solid",start_color=LIGHT_GRAY)
    ws2.column_dimensions["A"].width=22; ws2.column_dimensions["B"].width=3
    for col in ["C","D","E","F","G"]: ws2.column_dimensions[col].width=20

    # Hoja 3: Detalle
    ws3=wb.create_sheet("Detalle Partidas"); ws3.sheet_properties.tabColor=ACC
    ws3.merge_cells("A1:K1"); ws3["A1"]=f"DETALLE  -  {header.get('razon_social','').upper()}"
    hdr(ws3["A1"],bg=ACC,size=12); ws3.row_dimensions[1].height=28
    dh=["#","No. Articulo","Descripcion","Cant.","U.M.","Precio Unit.","Importe","Desc.","IVA 16%","Total c/IVA","Clave SAT"]
    dw=[5,13,52,7,6,15,15,12,12,14,12]
    for ci,(h,w) in enumerate(zip(dh,dw),1):
        c=ws3.cell(row=2,column=ci,value=h); hdr(c,bg=ACC2,size=9)
        ws3.column_dimensions[get_column_letter(ci)].width=w
    ws3.row_dimensions[2].height=32
    DR=3
    for idx,it in enumerate(items,1):
        r=DR+idx-1; bg=LIGHT_GRAY if idx%2==0 else None
        vals=[idx,it["no_identif"],it["descripcion"],it["cantidad"],it["um"],
              it["precio_unit"],f"=D{r}*F{r}",it["descuento"],f"=G{r}*0.16",f"=G{r}+I{r}",it.get("clave_sat","")]
        for ci,val in enumerate(vals,1):
            c=ws3.cell(row=r,column=ci,value=val); c.font=Font(name="Arial",size=9); c.border=_thin()
            if bg: c.fill=PatternFill("solid",start_color=bg)
            if ci in (1,4,5,11): c.alignment=Alignment(horizontal="center",vertical="center")
            elif ci in (6,7,8,9,10): c.alignment=Alignment(horizontal="right",vertical="center"); c.number_format="#,##0.00"
            else: c.alignment=Alignment(horizontal="left",vertical="center")
    ld=DR+len(items)-1
    def tr_row(ws,r,label,formula,bgl=ACC2,bgv="BDD7EE",bold=False,size=10):
        ws.merge_cells(f"A{r}:F{r}"); ws[f"A{r}"]=label; hdr(ws[f"A{r}"],bg=bgl,size=size,center=False)
        ws[f"G{r}"]=formula; ws[f"G{r}"].font=Font(name="Arial",bold=bold,size=size,color=GOLD if bold else "000000")
        ws[f"G{r}"].number_format='"$"#,##0.00'; ws[f"G{r}"].fill=PatternFill("solid",start_color=bgv)
        ws[f"G{r}"].alignment=Alignment(horizontal="right"); ws[f"G{r}"].border=_thin()
    t=ld+2
    tr_row(ws3,t,"SUBTOTAL (sin IVA)",f"=SUM(G{DR}:G{ld})")
    tr_row(ws3,t+1,"IVA 16%",f"=SUM(I{DR}:I{ld})")
    tr_row(ws3,t+2,"TOTAL",f"=G{t}+G{t+1}",bgl=ACC,bgv=ACC,bold=True,size=12)
    ws3.freeze_panes="A3"

    # Hoja 4: Resumen
    ws4=wb.create_sheet("Resumen"); ws4.sheet_properties.tabColor=ORANGE
    ws4.merge_cells("A1:F1"); ws4["A1"]=f"RESUMEN  -  {header.get('razon_social','').upper()}"
    hdr(ws4["A1"],bg=ACC,size=13); ws4.row_dimensions[1].height=30
    kpis=[("TOTAL PARTIDAS",len(items),None,ACC2),("PRODUCTOS UNICOS",len(unique),None,GREEN),
          ("SUBTOTAL",totals["subtotal"],'"$"#,##0.00',ACC2),("IVA 16%",totals["iva"],'"$"#,##0.00',ORANGE),
          ("TOTAL",totals["total"],'"$"#,##0.00',ACC),
          ("MULT. PUBLICO",f"x {mult_pub}",None,"1E7A4A"),("MULT. MAYOREO",f"x {mult_may}",None,"1F4E8A")]
    for i,(label,val,fmt,col) in enumerate(kpis):
        r=3+i
        ws4.merge_cells(f"A{r}:B{r}"); ws4[f"A{r}"]=label
        ws4[f"A{r}"].font=Font(name="Arial",bold=True,size=10,color=WHITE)
        ws4[f"A{r}"].fill=PatternFill("solid",start_color=col)
        ws4[f"A{r}"].alignment=Alignment(horizontal="left",vertical="center"); ws4[f"A{r}"].border=_thin()
        ws4.merge_cells(f"C{r}:F{r}"); ws4[f"C{r}"]=val
        ws4[f"C{r}"].font=Font(name="Arial",bold=True,size=11)
        ws4[f"C{r}"].fill=PatternFill("solid",start_color=LIGHT_GRAY)
        ws4[f"C{r}"].alignment=Alignment(horizontal="right" if fmt else "left",vertical="center"); ws4[f"C{r}"].border=_thin()
        if fmt: ws4[f"C{r}"].number_format=fmt
    for col in ["A","B","C","D","E","F"]: ws4.column_dimensions[col].width=22

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()

# ── Endpoints ─────────────────────────────────────────────────────────────────
@app.route("/generar-erp", methods=["POST"])
def generar_erp():
    if "pdf" not in request.files: return jsonify({"error":"No se recibio PDF"}),400
    f=request.files["pdf"]
    if not f.filename.lower().endswith(".pdf"): return jsonify({"error":"Solo PDFs"}),400
    empresa=request.form.get("empresa","vifer").lower()
    if empresa not in ("vifer","alessia"): return jsonify({"error":"Empresa invalida"}),400
    try: mp=float(request.form.get("mult_publico",2.3)); mm=float(request.form.get("mult_mayoreo",1.35))
    except: return jsonify({"error":"Multiplicadores invalidos"}),400
    try:
        pdf=f.read()
        data=parse_alessia(pdf) if empresa=="alessia" else parse_cfdi(pdf)
        if not data["items"]: return jsonify({"error":"No se encontraron partidas"}),422
        xlsx=build_excel(data,mp,mm)
        doc=(data["header"].get("no_pedido_sap") or data["header"].get("no_factura") or "DOC").replace(" ","")
        return send_file(io.BytesIO(xlsx),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,download_name=f"ERP_{empresa.upper()}_{doc}.xlsx")
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"error":str(e)}),500

@app.route("/health")
def health(): return jsonify({"status":"ok"})

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port)